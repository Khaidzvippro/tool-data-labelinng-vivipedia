"""
excel_writer.py — Tạo và ghi vào file Excel annotation v10.

- Không copy tem.xlsx — tạo workbook từ đầu nếu chưa có
- Append rows, không tạo file mới mỗi lần chạy
- Format: 4 dòng header, freeze pane C5, màu FF-prefix
"""
import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "outputs", "annotation_output.xlsx")

# ── Màu fill — BẮT BUỘC dùng prefix FF (8 ký tự) ────────────────────────────
# openpyxl hiểu 6 ký tự hex là alpha=00 (trong suốt) → màu không hiện
_NAVY = PatternFill("solid", fgColor="FF1F3864")   # IDENTITY (A-F)
_RED  = PatternFill("solid", fgColor="FFC00000")   # FACT-CHECK (G, H)
_BLUE = PatternFill("solid", fgColor="FF2E75B6")   # METRICS (I-L)
_GRAY = PatternFill("solid", fgColor="FF374151")   # ANNOTATION INFO (M-O)
_DATABG  = PatternFill("solid", fgColor="FFEBF8FF") # nền data rows
_STTBG   = PatternFill("solid", fgColor="FFF2F2F2") # nền cột STT

# ── Độ rộng cột A-O ───────────────────────────────────────────────────────────
_COL_WIDTHS = [5, 20, 16, 20, 12, 48, 18, 32, 9, 9, 9, 9, 40, 11, 14]

# ── Nhãn cột ──────────────────────────────────────────────────────────────────
_COL_LABELS = [
    "#",
    "Article / Page Title",
    "Domain",
    "Sub-domain",
    "Sub-domain\nID",
    "Claim (block nguyên văn)",
    "Fact-check\nStatus",
    "Fact-check\nSource URL",
    "Source\nFidelity\n(SF)",
    "Source\nCoverage\n(SC)",
    "Hallucination\nRate (HR)\n(inv.)",
    "Source\nQuality\n(SQ)",
    "Annotator Notes",
    "Annotator\nID",
    "Date",
]

_TEMPLATE_VALS = [
    "auto",
    "[article title]",
    "[select domain]",
    "[select sub-domain]",
    "[sub_id]",
    "[paste claim block]",
    "[XAC NHAN / LECH / MAU THUAN / OUTDATED / KHONG TIM THAY / BO QUA]",
    "[https://...]",
    "0-1", "0-1", "0-1", "0-1",
    "[SF=... / SC=... / HR=... / SQ=... / RISK=... / TXT=...]",
    "[ANT-xx]",
    "[date]",
]


def _col_fill(col: int) -> PatternFill:
    if col in {7, 8}:          return _RED
    if col in {9, 10, 11, 12}: return _BLUE
    if col in {13, 14, 15}:    return _GRAY
    return _NAVY


def _hdr(cell, value: str, fill: PatternFill, size: int = 8):
    cell.value     = value
    cell.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=size)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _data(cell, value, bold=False, size=9, color="FF000000",
          fill: PatternFill | None = None, h_align="left"):
    cell.value     = value
    cell.font      = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=h_align, vertical="top", wrap_text=True)
    if fill:
        cell.fill = fill


# ─────────────────────────────────────────────────────────────────────────────

def _build_header(ws):
    """Tạo 4 dòng header cho sheet Annotation."""
    # Row 1 — title bar
    ws.merge_cells("A1:O1")
    _hdr(ws["A1"],
         "RAG ANNOTATION TEMPLATE  —  SF / SC / HR / SQ per-claim  |  Rel / Comp cấp bài",
         _NAVY, size=11)
    ws.row_dimensions[1].height = 18

    # Row 2 — group labels
    ws.merge_cells("A2:F2")
    ws.merge_cells("G2:H2")
    ws.merge_cells("I2:L2")
    ws.merge_cells("M2:O2")
    _hdr(ws["A2"], "IDENTITY",              _NAVY, size=9)
    _hdr(ws["G2"], "FACT-CHECK",            _RED,  size=9)
    _hdr(ws["I2"], "METRICS (4 per-claim)", _BLUE, size=9)
    _hdr(ws["M2"], "ANNOTATION INFO",       _GRAY, size=9)
    ws.row_dimensions[2].height = 18

    # Row 3 — column labels
    for col, label in enumerate(_COL_LABELS, 1):
        _hdr(ws.cell(row=3, column=col), label, _col_fill(col), size=8)
    ws.row_dimensions[3].height = 32

    # Row 4 — template/instruction
    for col, val in enumerate(_TEMPLATE_VALS, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font      = Font(name="Arial", size=7, italic=True, color="FF888888")
        cell.fill      = _DATABG if col > 1 else _STTBG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Column widths
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze pane
    ws.freeze_panes = "C5"


def get_or_create_workbook() -> tuple[openpyxl.Workbook, str]:
    """
    Trả về (workbook, path).
    Nếu file chưa tồn tại → tạo mới với 4-row header.
    Nếu đã có → load lên để append.
    """
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    if os.path.exists(OUTPUT_PATH):
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        if "Annotation" not in wb.sheetnames:
            ws = wb.create_sheet("Annotation")
            _build_header(ws)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Annotation"
        _build_header(wb.active)
    return wb, OUTPUT_PATH


def _next_stt(ws) -> int:
    """STT tiếp theo — bỏ qua 4 dòng header, data từ row 5."""
    for r in range(ws.max_row, 4, -1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            try:
                return int(val) + 1
            except (TypeError, ValueError):
                pass
    return 1


def append_rows(rows: list[list], log_fn=print) -> str:
    """
    Append danh sách hàng vào sheet Annotation.
    rows: list of list, mỗi list có 15 phần tử [A..O].
          Phần tử A (STT) truyền "" → hàm tự gán số.
    Trả về đường dẫn tuyệt đối file Excel.
    """
    wb, path = get_or_create_workbook()
    ws = wb["Annotation"]

    # Tìm dòng bắt đầu append
    start_row = max(ws.max_row + 1, 5)
    # Nếu dòng cuối là dòng trống (template row 4 chưa có data)
    if ws.max_row >= 5 and ws.cell(row=ws.max_row, column=1).value is None:
        start_row = ws.max_row

    stt = _next_stt(ws)

    for i, row_data in enumerate(rows):
        r = start_row + i

        # Gán STT vào cột A
        data_row = list(row_data)
        data_row[0] = stt + i

        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=r, column=col)
            bg   = _STTBG if col == 1 else _DATABG

            if col in {9, 10, 11, 12}:   # SF/SC/HR/SQ — bold, center
                _data(cell, val, bold=True, size=10, fill=bg, h_align="center")
            elif col == 8:                # URL — màu xanh link
                _data(cell, val, size=8, color="FF0563C1", fill=bg)
            elif col == 1:                # STT — xám, center
                _data(cell, val, size=9, color="FF888888", fill=bg, h_align="center")
            else:
                _data(cell, val, fill=bg)

        ws.row_dimensions[r].height = 200

    try:
        wb.save(path)
    except PermissionError:
        import time
        alt = path.replace(".xlsx", f"_{int(time.time())}.xlsx")
        wb.save(alt)
        path = alt
        log_fn(f"  ⚠ File gốc đang mở → đã lưu vào: {alt}")

    log_fn(f"  Đã ghi {len(rows)} hàng vào: {path}")
    return os.path.abspath(path)


# Backward compat — main cũ gọi write_output
def write_output(stt: str, data: dict, log_fn=print) -> str:
    """Legacy wrapper — chuyển data Claude → rows và append."""
    art    = data.get("article", {})
    claims = data.get("claims", [])
    today  = date.today().strftime("%Y-%m-%d")

    rows = []
    for c in claims:
        rows.append([
            "",
            art.get("title", ""),
            art.get("domain", ""),
            art.get("sub_domain", ""),
            art.get("sub_domain_id", ""),
            c.get("claim", ""),
            c.get("fact_check_status", ""),
            c.get("fact_check_source_url", ""),
            c.get("source_fidelity", ""),
            c.get("source_coverage", ""),
            c.get("hallucination_rate", ""),
            c.get("source_quality", ""),
            c.get("notes", ""),
            "AUTO",
            today,
        ])
    return append_rows(rows, log_fn=log_fn)
